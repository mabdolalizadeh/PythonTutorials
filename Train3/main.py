import numpy as np
import matplotlib.pyplot as plt
from openpyxl import *

workSheet = load_workbook("data.xlsx")["Sheet1"]
firstCloumn = []
secondCloumn = []
for i in range(1, workSheet.max_row + 1):
    firstCloumn.append(workSheet[f"A{i}"].value)

for j in range(1, workSheet.max_row + 1):
    secondCloumn.append(workSheet[f"B{j}"].value)

x = np.array(firstCloumn)
y = np.array(secondCloumn)
dy = y[1:] - y[:-1]
dx = x[1:] - x[:-1]

d2y = dy[1:] - dy[:-1]

diff1 = dy/dx
diff2 = d2y/dx[1:]



